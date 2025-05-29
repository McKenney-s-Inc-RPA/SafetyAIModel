from __future__ import annotations
import pandas as pd
from overall_helpers import *
from data_collection import DataCollection, get_base_file_name_without_ext
from typing import Callable
from openpyxl import load_workbook
from os.path import exists as path_exists, join as path_join, getmtime as file_getmtime
from os import mkdir, remove as remove_file
import pantab
import warnings
from enum import Enum


class CSVParserOptions:
    def __init__(self):
        self.date_columns: list[str | int] = []
        self.column_types: dict | str | None = None


class DataTable(DataCollection):
    # a representation of a file, containing >=1 dimensions (i.e., independent variables)
    # and >=1 measures (i.e., dependent variables)
    # these dimensions can be common between multiple files
    # also contains the actual data set itself (loaded into pandas for ease of use)
    def __init__(self, file_name: str, auto_populate: bool = True, max_rows_to_sample: int | None = None,
                 sheet_name: int | str | None = 0, parserOptions: CSVParserOptions = None):
        self._file_name = file_name
        display_name = DataCollection.get_file_display_name(file_name, sheet_name)
        diagnostic_log_message(f"Reading in data for {display_name}")
        if file_name.endswith(".xlsx"):
            df = pd.read_excel(file_name, sheet_name=sheet_name, header=0, nrows=max_rows_to_sample)
        elif file_name.endswith(".csv"):
            fileDtypes = None
            parse_dates = False
            if parserOptions is not None:
                fileDtypes = parserOptions.column_types
                if len(parserOptions.date_columns) > 0:
                    parse_dates = parserOptions.date_columns
            df = pd.read_csv(file_name, header=0, nrows=max_rows_to_sample, dtype=fileDtypes, parse_dates=parse_dates)
        elif file_name.endswith(".hyper"):
            table_name = get_base_file_name_without_ext(file_name) if sheet_name is None or sheet_name == 0 else sheet_name
            df = pantab.frame_from_hyper(file_name, table=table_name)
        else:
            raise ValueError("Unrecognized file type given")

        # import the data frame, and show the headers
        super().__init__(df, display_name)
        diagnostic_log_message(f"{display_name} has headers: {self.headers}")
        if auto_populate:
            self.auto_assign_measures()


class CertaintyOptions(Enum):
    UNKNOWN = -1
    NO = 0
    YES = 1


class DataFolder:
    def __init__(self, name: str, parent: DataFolder | None):
        self.parent = parent
        self.name = name
        self._full_path: str = name
        self.data_files: Dict[str, DataFile] = {}
        self.children: Dict[str, DataFolder] = {}
        self._exists = CertaintyOptions.UNKNOWN
        self._contains_any_data_sources = CertaintyOptions.UNKNOWN

        if parent is not None:
            parent.children[name] = self
            self._full_path = parent.get_sub_path(name)

    def get_sub_path(self, sub_name):
        return path_join(self._full_path, sub_name)

    @property
    def full_path(self):
        return self._full_path

    @property
    def exists(self) -> bool:
        if self._exists == CertaintyOptions.UNKNOWN:
            # if is unknown, then find out
            exists_check: bool
            if self.parent is not None and not self.parent.exists:
                # if the parent is set, but doesn't exist, then return false
                exists_check = False
            else:
                # since is the root, or if the parent exists, then check if the folder exists
                exists_check = path_exists(self._full_path)
            self._exists = CertaintyOptions.YES if exists_check else CertaintyOptions.NO
        return self._exists == CertaintyOptions.YES

    @property
    def contains_any_data_sources(self) -> bool:
        if self._contains_any_data_sources == CertaintyOptions.UNKNOWN:
            # if this contains any data file, then return true
            if len(self.data_files) > 0:
                self._contains_any_data_sources = CertaintyOptions.YES
                return True
            # next, check if any children contain data files
            for c in self.children.values():
                if c.contains_any_data_sources:
                    self._contains_any_data_sources = CertaintyOptions.YES
                    return True
            self._contains_any_data_sources = CertaintyOptions.NO
            return False
        return self._contains_any_data_sources == CertaintyOptions.YES

    def _create_if_necessary(self):
        # only perform if sure that doesn't exist yet
        if not self.exists:
            # if the parent is set, then create if needed
            if self.parent is not None:
                self.parent._create_if_necessary()
            # now, make the current directory
            mkdir(self._full_path)
            self._exists = CertaintyOptions.YES

    def get_folder(self, folder_name: str):
        if folder_name in self.children:
            return self.children[folder_name]
        else:
            # don't create until add data sources to it
            return DataFolder(folder_name, self)

    def build_necessary_structure(self):
        # create if necessary
        if self.contains_any_data_sources:
            self._create_if_necessary()
            # build out any data sources
            for d in self.data_files.values():
                d.build_files()
        for c in self.children.values():
            c.build_necessary_structure()

    def get_data_source(self, source_name: str, debug_mode: bool = DEBUG_DEFAULT, max_rows_to_sample: int | None = None,
                        creator_func: Callable[[], DataCollection] | List[str] | None = None) -> DataFile:
        # the source name will NOT contain an extension
        # the source name will either be saved as a .hyper file or a .csv file
        # if the debug mode is true, then will convert over to a .csv file
        # if the debug mode is false, then will convert over to a .hyper file
        # the creator function is what creates the data collection if doesn't already exist
        if source_name in self.data_files:
            return self.data_files[source_name]
        else:
            new_file = DataFile(source_name, self, creator_func, max_rows_to_sample)
            new_file.debug_mode = debug_mode
            new_file.save_after_create = CACHE_AS_PROCESS
            self.data_files[source_name] = new_file
            return new_file

    def save_data(self, source_name: str, new_data: DataCollection, debug_mode: bool = DEBUG_DEFAULT, save_now=False):
        file_val = self.get_data_source(source_name, debug_mode=debug_mode)
        file_val.set_data(new_data, build_files=save_now)
        return file_val


def is_modified_before(file_name: str, modified_threshold: datetime):
    if path_exists(file_name):
        modified_date_float = file_getmtime(file_name)
        modified_date = datetime.fromtimestamp(modified_date_float)
        return modified_date < modified_threshold
    return False


def get_modified_date(file_name: str, fallback_modified_date: datetime | None = None):
    if path_exists(file_name):
        modified_date_float = file_getmtime(file_name)
        return datetime.fromtimestamp(modified_date_float)
    elif fallback_modified_date is not None:
        return fallback_modified_date
    else:
        return datetime(1970, 1, 1)


class DataFile:
    # information about a data table, for ease of saving
    # note that only loads / saves when needed, which should save time
    # the source name will NOT contain an extension
    # the source name will be saved as a .hyper file or a .csv file, or both
    # if the debug mode is true, then will convert over to a .csv file
    # if the debug mode is false, then will convert over to a .hyper file
    # the creator function is what creates the data collection if doesn't already exist
    def __init__(self, source_name: str, parent: DataFolder,
                 creator_func: Callable[[], DataCollection] | List[str] | None = None,
                 max_rows_to_sample: int | None = None):
        if source_name.endswith(".csv"):
            source_name = source_name[:-4]
        elif source_name.endswith(".hyper"):
            source_name = source_name[:-6]

        self._source_name = source_name
        self._parent = parent
        self._creator_func = creator_func
        self.max_rows_to_sample = max_rows_to_sample
        self._hyper_source = parent.get_sub_path(source_name + ".hyper")
        self._csv_source = parent.get_sub_path(source_name + ".csv")
        # rebuild the file if want to rebuild files and the creator function is provided, and was created before the cached file cutoff
        rebuild_file = False
        if REBUILD_CACHED_FILES and self._creator_func is not None:
            rebuild_file = is_modified_before(self._hyper_source, REBUILD_CACHED_FILES_BEFORE) or \
                           is_modified_before(self._csv_source, REBUILD_CACHED_FILES_BEFORE)
        # use cached file if don't want to rebuild
        self._use_cached = not rebuild_file
        self._loaded_data: DataCollection | None = None
        self._debug_mode = False
        self._save_csv = False
        self._save_hyper = False
        self._save_both = False
        self._data_has_changed = False
        self.save_after_create = False

    @property
    def use_cached(self) -> bool:
        return self._use_cached

    @use_cached.setter
    def use_cached(self, value: bool):
        # if want to use cached value, then set as true
        # note that if already is loaded, then don't have to throw away
        # however, if don't want to use cached value, and has a way to create the value, then set back to a null value
        if not value and self._use_cached and self._loaded_data is not None:
            if self._creator_func is None:
                warnings.warn(f"Throwing away the cached source for '{self._source_name}' without a creator function could cause problems...")
            self._loaded_data = None
        self._use_cached = value

    def _set_csv_and_hyper_settings(self):
        # if in debug mode, then will want to see the csv file
        # otherwise, will want to see the hyper file
        # note that if the user wants to save both, then don't perform
        self._save_csv = self._save_both or self._debug_mode
        self._save_hyper = self._save_both or not self._debug_mode

    @property
    def debug_mode(self):
        return self._debug_mode

    @debug_mode.setter
    def debug_mode(self, value: bool):
        self._debug_mode = value
        self._set_csv_and_hyper_settings()

    @property
    def save_both(self):
        return self._save_both

    @save_both.setter
    def save_both(self, value: bool):
        self._save_both = value
        self._set_csv_and_hyper_settings()

    @property
    def exists(self) -> bool:
        return path_exists(self._hyper_source) or path_exists(self._csv_source)

    def get_data(self, auto_populate=False, name: str | None = None, parserOptions: CSVParserOptions = None) -> DataCollection:
        # returns the data collection, and loads / creates if needed
        if self._loaded_data is None:
            # check to see if the file(s) exist
            if self._use_cached and self._parent.exists:
                # first, check to see if the hyper file exists
                # otherwise, check to see if the csv file exists
                # for cached files, don't autopopulate, as should already have all data loaded
                if path_exists(self._hyper_source):
                    self._loaded_data = DataTable(self._hyper_source, auto_populate=auto_populate, max_rows_to_sample=self.max_rows_to_sample)
                elif path_exists(self._csv_source):
                    self._loaded_data = DataTable(self._csv_source, auto_populate=auto_populate, max_rows_to_sample=self.max_rows_to_sample,
                                                  parserOptions=parserOptions)
                # if the data is still not loaded, then create
                if self._loaded_data is None:
                    if self._creator_func is not None:
                        # run the creator function
                        self._set_data(self._creator_func())
                        # if file already exists, then mark as having changed (which will save over it)
                        self._data_has_changed = self.exists or self._use_cached
                        # if want to save, then do so
                        if self.save_after_create:
                            self.build_files()
                    else:
                        raise NotImplementedError(f"No creator function provided for '{self._source_name}'")
        # create a copy, so that if modifies the data collection, then doesn't save the modifications
        # could happen if are swapping from one file format to another, even if the data isn't explicitly written over.
        return self._loaded_data.copy(name)

    def set_data(self, new_data, build_files=False):
        if new_data != self._loaded_data:
            self._set_data(new_data)
            self._data_has_changed = True
            if build_files:
                self.build_files()

    def _set_data(self, new_data):
        # note that may be a data collection or a pandas data frame
        if isinstance(new_data, pd.DataFrame):
            # convert over to a data collection
            self._loaded_data = DataCollection(new_data, self._source_name)
        else:
            self._loaded_data = new_data

    def _save_or_delete_file(self, file_name: str, should_save: bool, file_exists_flag: bool | None = None):
        if should_save:
            self._loaded_data.save_data(file_name)
        else:
            if file_exists_flag is None:
                file_exists_flag = path_exists(file_name)
            if file_exists_flag:
                remove_file(file_name)

    def build_files(self):
        # builds the files
        # if the data has changed since loaded, then should save the files
        # otherwise, if the data hasn't changed, then check to see if the appropriate files are written
        # if they are not, then write as needed
        if self._data_has_changed:
            # since it has changed, delete out any existing files, and rewrite
            self._save_or_delete_file(self._csv_source, self._save_csv)
            self._save_or_delete_file(self._hyper_source, self._save_hyper)
        else:
            # since has not changed, see if need to resave the files in a different format
            csv_file_exists = path_exists(self._csv_source)
            hyper_file_exists = path_exists(self._hyper_source)
            # will need to load if one of the files is wanting to be saved, but doesn't exist yet
            need_to_load_data = (self._save_csv and not csv_file_exists) or (self._save_hyper and not hyper_file_exists)
            if need_to_load_data:
                self.get_data()
            # now, resave / delete only if doesn't match the current flags
            if csv_file_exists != self._save_csv:
                self._save_or_delete_file(self._csv_source, self._save_csv, csv_file_exists)
            if hyper_file_exists != self._save_hyper:
                self._save_or_delete_file(self._hyper_source, self._save_hyper, hyper_file_exists)


def source_exists(file_name: str, sheet_name: str | None = None):
    # check if the file exists
    if not path_exists(file_name):
        return False
    # if the sheet name is given and the file name is an excel file, then check it
    if sheet_name is not None and file_name.endswith(".xlsx"):
        # load the file and make sure that the sheet exists
        wb = load_workbook(file_name, read_only=True)
        return sheet_name in wb.sheetnames
    else:
        return True


def load_or_create_data(file_name: str, creator_func: Callable[[], DataCollection | pd.DataFrame] | List[str],
                        auto_populate: bool = True, max_rows_to_sample: int | None = None,
                        sheet_name: int | str | None = 0, modify_func: Callable[[DataCollection], None] | None = None,
                        use_cached_file=True, save_cached_file=True) -> DataCollection:
    # if the file exists and should use the cached file, then return the value
    if source_exists(file_name, sheet_name) and use_cached_file:
        diagnostic_log_message(f"Reading {DataCollection.get_file_display_name(file_name, sheet_name)} from cache...")
        read_file = DataTable(file_name, auto_populate=auto_populate, max_rows_to_sample=max_rows_to_sample,
                              sheet_name=sheet_name)
        if modify_func is not None:
            modify_func(read_file)
        return read_file
    else:
        diagnostic_log_message(f"{DataCollection.get_file_display_name(file_name, sheet_name)} not found. Creating data instead...")
        # otherwise, create the output. Note that could be a dataframe or a data collection.
        if isinstance(creator_func, list):
            # should be a list of headers for a new data file
            created_file = pd.DataFrame(columns=creator_func)
        else:
            created_file = creator_func()
        created_collection: DataCollection = created_file
        if isinstance(created_file, pd.DataFrame):
            # convert over to a data collection
            created_collection = DataCollection(created_file, DataCollection.get_file_display_name(file_name,
                                                                                                   sheet_name))
        if save_cached_file:
            created_collection.save_data(file_name, sheet_name)
        return created_collection
